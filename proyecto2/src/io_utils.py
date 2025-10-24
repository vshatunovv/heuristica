
import json, shutil, os
import pandas as pd
from openpyxl import load_workbook

def load_instance(inst_dir: str):
    employees = pd.read_csv(os.path.join(inst_dir, "employees.csv"))
    desks = pd.read_csv(os.path.join(inst_dir, "desks.csv"))
    compat_path = os.path.join(inst_dir, "compatibility.csv")
    if os.path.exists(compat_path):
        compat = pd.read_csv(compat_path)
    else:
        compat = pd.DataFrame([(e, d, 1) for e in employees.employee_id for d in desks.desk_id],
                              columns=["employee_id","desk_id","compatible"])
    with open(os.path.join(inst_dir, "params.json"), "r") as f:
        params = json.load(f)
    return employees, desks, compat, params

def write_to_excel(plantilla_path: str, out_path: str, assignment: dict, group_days: dict, summary: dict):
    shutil.copyfile(plantilla_path, out_path)
    wb = load_workbook(out_path)
    wsA = wb["EmployeeAssignment"]
    wsG = wb["Groups Meeting day"]
    wsS = wb["Summary"]

    wsA.delete_rows(2, wsA.max_row)
    for e, days in assignment.items():
        wsA.append([e, days.get("Mon","none"), days.get("Tue","none"), days.get("Wed","none"),
                    days.get("Thu","none"), days.get("Fri","none")])

    wsG.delete_rows(2, wsG.max_row)
    for g, d in group_days.items():
        wsG.append([g, d])

    wsS.delete_rows(2, wsS.max_row)
    for k,v in [("Valid assignments", summary["Valid assignments"]),
                ("Employee preferences", summary["Employee preferences"]),
                ("Isolated employees", summary["Isolated employees"])]:
        wsS.append([k, v])

    wb.save(out_path)
